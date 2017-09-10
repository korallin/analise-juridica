#!/usr/bin/python
# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import scipy.stats as stats
import statsmodels.api as sm
import re

import sys
from openpyxl import Workbook, load_workbook


file_name = sys.argv[1]
wb = load_workbook(file_name)

def calc_qui_square_test(wb, file_name, sheet_name, skiprows=None):
    df = pd.read_excel(file_name, sheet_name, index_col=0, header=0, skiprows=skiprows)
    quisq_test, p_value, dof, _ = stats.chi2_contingency(np.array([df["Alg 1"].values, df["Alg 2"].values]))

    print sheet_name, quisq_test, p_value, dof, _

    # write info to sheet
    ws = wb[sheet_name]
    ws['A' + str(ws.max_row + 1)] = ''
    ws.append(['Chi-square test', 'p value'])
    ws.append([quisq_test, p_value])

    file_name = re.search(r"(.*).xlsx", file_name).groups(1)[0] + "_post_processed.xlsx"
    wb.save(file_name)


# def calc_proportion_ztest(wb, file_name, sheet_name, index, nobs):
#     df = pd.read_excel(file_name, sheet_name, index_col=0, header=0)
#     count = np.array([df.loc[index, "Alg 2"], df.loc[index, "Alg 1"]])
#     zstat, p_value = sm.stats.proportions_ztest(count, nobs)

#     # write info to sheet
#     ws = wb[sheet_name]
#     ws['A' + str(ws.max_row + 1)] = 'row ' + index
#     ws.append(['prop ztest test', 'p value'])
#     ws.append([zstat, p_value])

#     wb.save(file_name + "post_processed")



skiprows_lst = [None, None, [3], None, None, [4]]
sheet_name_lst = ['min 8 top 100 page rank', 'min 9 top 100 page rank',
                'min 10 top 100 page rank',
                'min 8 top 100 page rank - ministros',
                'min 9 top 100 page rank - ministros',
                'min 10 top 100 page rank - ministros']
[calc_qui_square_test(wb, file_name, sheet_name, skiprows) for sheet_name, skiprows 
                                                in zip(sheet_name_lst, skiprows_lst)]

# linha 4 do sheet 'min 10 top 100 page rank - ministros'
# nobs = np.array([231, 237])
# calc_proportion_ztest(wb, file_name, 'min 10 top 100 page rank - ministros', "DIAS TOFFOLI", nobs)
# # sheet 'min 10 top 100 page rank' para cada linha (ou para a última que é a única que possui problema)
# nobs = np.array([223, 240])
# calc_proportion_ztest(wb, file_name, 'min 10 top 100 page rank', u"70% de decisões", nobs)

