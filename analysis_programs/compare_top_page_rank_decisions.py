#!/usr/bin/env python
# -*- coding: utf-8 -*-

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import matplotlib.cm as cm
import matplotlib.colors as col
from math import log, log10

from pymongo import MongoClient
from operator import itemgetter
from random import randint
import sys
import os
import numpy as np
import re

from openpyxl import Workbook, load_workbook

reload(sys)
sys.setdefaultencoding('utf8')

# python compare_top_page_rank_decisions.py DJTest stf_pr_1_acordaos
# python compare_top_page_rank_decisions.py DJTest stf_pr_2_acordaos
mongo_user = sys.argv[1]
mongo_password = sys.argv[2]
dbName = sys.argv[3]
collection_name = sys.argv[4]

spreadsheet_name = "page_ranker_analysis.xlsx"
if os.path.isfile(spreadsheet_name):
    wb = load_workbook(spreadsheet_name)
else:
    wb = Workbook()

ws = wb.create_sheet(collection_name)


# __file__ = '/home/jackson/analise-juridica/Scripts/'
# dbName = "DJTest"
# collection_name = "stf_pr_2_acordaos"


client = MongoClient('mongodb://{}:{}@127.0.0.1:57017'.format(mongo_user, mongo_password))
db = client[dbName]


def num_rel_decs_citadas_lst(relator_decs_citadas_dict, dec_ids):
    t_df_lst = []
    for dec_id in dec_ids:
        t_df = 0
        for citados_lst in relator_decs_citadas_dict.values():
            if dec_id in citados_lst:
                t_df += 1
        t_df_lst.append(t_df)

    return t_df_lst

def get_dir_path(dir):
    dir_path = os.path.dirname(os.path.abspath(__file__))
    graficos_dir = os.path.join(dir_path, dir)
    if not os.path.exists(graficos_dir):
        os.makedirs(graficos_dir)

    return graficos_dir


def generate_histogram(array, title, xlabel, ylabel, file_name, sim_descr):
    fig, ax = plt.subplots(1)

    result = ax.hist(array, color='c')

    mean = np.mean(array)
    median = np.median(array)

    plt.axhline(mean, color='r', linestyle='dashed', linewidth=2)
    plt.axhline(median, color='g', linestyle='dotted', linewidth=2)

    plt.title(title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)

    N = len(array)
    ax.set_xticks(N)
    xtickNames = ax.set_xticklabels(xtick_names)
    plt.setp(xtickNames, rotation=45, fontsize=10)
    plt.tight_layout()

    dir_path = "graficos_%s" % sim_descr
    dir_path = get_dir_path(dir_path)

    plt.savefig(dir_path + '/{}.png'.format(file_name))
    plt.close()


def generate_barplot(array, title, xlabel, ylabel, file_name, xtick_names, sim_descr):
    fig, ax = plt.subplots(1)

    n = len(array)
    index = np.arange(n)
    values = [randint(1, 100) for i in range(n)]

    cmap = cm.ScalarMappable(col.Normalize(0, 99), cm.spring)

    bar_width = 0.85
    bars = plt.bar(index, array, bar_width, alpha=0.5, color=cmap.to_rgba(values))

    mean = np.mean(array[~np.isnan(array)])
    median = np.median(array[~np.isnan(array)])

    props = dict(boxstyle='round', facecolor='wheat', alpha=0.5)
    textstr = '$\mu=%.2f$\n$\mathrm{median}=%.2f$\n' % (mean, median)
    # if there is nan values at array add string to say that xtick x_i is null
    # think of index of nan to get x_i names
    ax.text(0.95, 0.95, textstr, transform=ax.transAxes, fontsize=14,
            verticalalignment='top', horizontalalignment='right', bbox=props)

    plt.axhline(mean, color='r', linestyle='dashed', linewidth=2)
    plt.axhline(median, color='g', linestyle='dotted', linewidth=2)

    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)

    xticks_pos = [0.65*patch.get_width() + patch.get_xy()[0] for patch in bars]
    plt.xticks(xticks_pos, xtick_names, rotation=45, fontsize=8, ha='right')
    # ax.set_xticks(index)
    # ax.set_xticklabels(xtick_names, rotation=45, fontsize=8, ha='center')
    plt.tight_layout()

    dir_path = "graficos_%s" % sim_descr
    dir_path = get_dir_path(dir_path)

    plt.savefig(dir_path + '/{}.png'.format(file_name))
    plt.close()


def percentage_citations(cit, c):
    cit_len = len(cit)
    own_author_len = len(filter(lambda x: x in c, cit))
    other_author_len = cit_len - own_author_len
    return (100 * own_author_len/float(cit_len), 100 * other_author_len/float(cit_len))

# recebe uma lista de tuplas e retorna uma tupla com o valor médio de cada posição da tupla
def lst_tuples_mean(tuple_lst):
    t_len = len(tuple_lst[0])
    lst_len = len(tuple_lst)

    lst = []
    for i in range(t_len):
        lst.append(sum([t[i] for t in tuple_lst])/lst_len)

    return tuple(lst)


def virtual_dec(dec):
    if dec['virtual'] == True:
        return 1
    return 0


def get_columns_len_tup(page_ranks_iter, data_type):

    if len(page_ranks_iter) == 0:
        return ()

    if data_type is dict:
        dec_id_len = max([len(pr["acordaoId"]) for pr in page_ranks_iter])
        relator_len = max([len(pr["relator"]) for pr in page_ranks_iter])
    elif data_type is list:
        dec_id_len = max([len(pr[0]) for pr in page_ranks_iter])
        relator_len = max([len(pr[1]) for pr in page_ranks_iter])
    else:
        dec_id_len = max([len(k) for k, v in page_ranks_iter])
        relator_len = max([len(v[1]) for k, v in page_ranks_iter])

    # 3° parâmetro corresponde a tamanho do espaço usado para marcar atributo como virtual
    return (dec_id_len, relator_len, 1)


def construct_string(values, lengths):
    string = ""
    for v, l in zip(values, lengths):
        v_str = str(v.encode('utf-8'))
        string += " %s" % v_str + " " * (l - len(v)) + " |"

    return string

removed_decisions_freq = {}
removed_decisions_iters = []
page_ranks_iters = []

selected_keys = {"acordaoId": 1, "pageRank": 1, 'virtual': 1, 'relator': 1, 'citacoes': 1, 'citadoPor': 1}
for i in range(1, 11):
    coll_rank = collection_name + "_%d" % i
    coll_removed = collection_name + "_removed_%d" % i

    page_ranks_cursor = db[coll_rank].find({}, selected_keys).sort([("pageRank", -1)]).limit(100)
    page_ranks_iters.append(list(page_ranks_cursor))

    removed_decisions_cursor = db[coll_removed].find({})
    rem_decs = list(removed_decisions_cursor)

    if rem_decs != []:
        removed_decisions_iters.append(rem_decs[0]['removed_decisions'])
        for rd in rem_decs[0]['removed_decisions']:
            if rd in removed_decisions_freq:
                removed_decisions_freq[rd] += 1
            else:
                removed_decisions_freq[rd] = 1
    else:
        removed_decisions_iters.append([])


# DESCRIÇÃO GERAL antes da por iteração
page_rank_ids_relatores_lists = [[(pr['acordaoId'], pr['relator'], pr['virtual']) for pr in pr_list] for pr_list in page_ranks_iters]

print "Descrição de informações da simulação do page rank por iteração\n"
# DESCRIÇÃO POR ITERAÇÃO
decisoes_removidas_no_top100 = []
virtual_decs = []
rel_freq_iter_dict = {}
rel_freqs_iter = []
for i in range(10):

    columns = get_columns_len_tup(decisoes_removidas_no_top100, list)

    print "Iteração %d" % (i+1)
    print "IDs presentes na iteraçao anterior e que foram removidos na simulação da iteração atual"
    print "DECISION ID | RELATOR | VIRTUAL"
    for pr_it in decisoes_removidas_no_top100:
        virtual = "S" if pr_it[2] is True else "N"
        string = construct_string([pr_it[0], pr_it[1], virtual], columns)
        print string[:-2]
    print ""

    virtual_decs.append(sum(map(virtual_dec, page_ranks_iters[i])))
    rel_freq_dict = {}

    columns = get_columns_len_tup(page_ranks_iters[i], dict)
    print "RANK | DECISION ID | RELATOR | VIRTUAL | PAGE RANK"
    for j, pr_it in enumerate(page_ranks_iters[i]):
        virtual = "S" if pr_it['virtual'] is True else "N"
        string = construct_string([pr_it['acordaoId'], pr_it['relator'], virtual], columns)
        string = " {}{} |".format(j+1, " " * (4 - len(str(j+1)))) + string + " {0:.6f}".format(pr_it['pageRank'])
        print string

        relator = pr_it['relator']
        if relator in rel_freq_dict:
            rel_freq_dict[relator] += 1
        else:
            rel_freq_dict[relator] = 1

        if relator in rel_freq_iter_dict:
            rel_it_dict = rel_freq_iter_dict[relator]
            if pr_it['acordaoId'] in rel_it_dict:
                rel_it_dict[pr_it['acordaoId']][0] += 1
            else:
                rel_it_dict[pr_it['acordaoId']] = [1, pr_it['citacoes'], pr_it['citadoPor']]
        else:
            rel_freq_iter_dict[relator] = {pr_it['acordaoId']: [1, pr_it['citacoes'], pr_it['citadoPor']]}


    rel_freqs_iter.append(sorted(rel_freq_dict.items(), key=itemgetter(1), reverse=True))

    if (i+1) < 10:
        decisoes_removidas_no_top100 = filter(lambda v: v[0] in removed_decisions_iters[i+1], page_rank_ids_relatores_lists[i])

    print "\n"

virtual_decs_a = np.array(virtual_decs)
print "Virtual decisions in each iteration"
for i, vd_number in enumerate(virtual_decs_a):
    print "%d: %d" % (i, vd_number)

print "Mean of virtual decisions: %.2f" % np.mean(virtual_decs_a)
print "Standard deviation of virtual decisions: %.2f" % np.std(virtual_decs_a)
print "\n"
ws.append(["Mean of virtual decisions", np.mean(virtual_decs_a)])
ws.append(["Standard deviation of virtual decisions", np.std(virtual_decs_a)])


rel_freqs_dict = {}
print "Exibindo frequência com que aparece cada relator em cada iteração de ordenado por frequência de ocorrência"
for i in range(10):
    print "\nIteração %d | # de relatores na iteração %d" % (i+1, len(rel_freqs_iter[i]))
    for r, f in rel_freqs_iter[i]:
        print "{}: {}".format(r.encode('utf-8'), f)
        if r in rel_freqs_dict:
            rel_freqs_dict[r].append(f)
        else:
            rel_freqs_dict[r] = [f]


relatores = []
relat_freqs = []
print "\nRelator | freq absol | iterações presente | média em iterações presente | std"
for (r, f_lst) in sorted(rel_freqs_dict.items(), key=lambda i: sum(i[1]), reverse=True):
    f_lst_array = np.array(f_lst)
    print "{} | {} ({:.2f}%) | {} | {:.2f} | {:.2f}".format(r.encode('utf-8'), sum(f_lst), 100 * sum(f_lst) / 1000., len(f_lst), np.mean(f_lst_array), np.std(f_lst_array))
    relatores.append(r.encode('utf-8'))
    relat_freqs.append(sum(f_lst))

sim_descr = re.search('(\d.+)', collection_name).groups()[0]

xlabel = 'Ministros'
ylabel = 'Número de decisões'
title = "Número de ocorrência de decisões por ministro \ncom repetição de ocorrência de decisões (%s)" % sim_descr
file_name = "ocorrências_de_ministros_com_repet_%s" % sim_descr
generate_barplot(np.array(relat_freqs), title, xlabel, ylabel, file_name, relatores, sim_descr)
# generate_histogram(np.array(relat_freqs), title, xlabel, ylabel, file_name, xtick_names=np.array(relatores))

relator_decs_citadas_dict = {}
for k, v in rel_freq_iter_dict.items():
    relator_decs_citadas_dict[k] = sum([cit[1] for cit in v.values()], [])

# soma tf-idf de decisões do relator (geral) | soma tf-idf de decisões do relator (top 100)
# cria dicionário de decisões citadas por cada relator no top 100 | obtém decisões citadas por relator no BD -> criar função que checa # de relatores que citam uma dada decisão
print "\nRelator | freq dec únicas | (freq dec únicas)/(freq absol) | (freq dec únicas)/(# decs relator) | (freq absol)/(# decs relator) | soma tf-idf de decisões do relator"
ws['A' + str(ws.max_row + 2)] = ''
ws.append(["Relator", "freq dec únicas", "(freq dec únicas)/(freq absol)",
            "(freq dec únicas)/(# decs relator)", "(freq absol)/(# decs relator)",
            "soma tf-idf de decisões do relator"])

relatores = []
relat_dec_num = []
citacoes_top100 =[]
porc_cit_mesmo_rel = []
porc_cit_por_mesmo_rel = []
coll_all_ids = collection_name + "_%d" % 1
for (r, f) in sorted(rel_freq_iter_dict.items(), key=lambda i: len(i[1]), reverse=True):
    dec_ids = f.keys()
    citacoes = [f[k] for k in dec_ids]
    relat_dec_num.append(len(citacoes))

    f_unica = len(citacoes)
    f_absol = sum(rel_freqs_dict[r])

    r_ids = [obj['acordaoId'] for obj in list(db[coll_all_ids].find({'relator': r}, {'acordaoId':1}))]

    N = float(len(r_ids))
    N_rel = len(rel_freq_iter_dict)
    # N_top100 = f_unica

    t_fd_lst = num_rel_decs_citadas_lst(relator_decs_citadas_dict, dec_ids)
    sum_tf_idf = sum([(1 + log(1 + len(cit[2]))) * log10(N_rel/t_fd) if t_fd > 0 else 0 for cit, t_fd in zip(citacoes, t_fd_lst)])

    print "{} | {} ({:.2f}%) | {:.2f} | {:.4f} | {:.4f} | {:.4f} ".format(r.encode('utf-8'), len(f),
                                                                            100 * len(f) / 1000.,
                                                                            f_unica/float(f_absol), f_unica/N,
                                                                            f_absol/N, sum_tf_idf)
    ws.append([r.encode('utf-8'), len(f), f_unica/float(f_absol), f_unica/N, f_absol/N, sum_tf_idf])
    # listas de tuplas com porcentagem de citações a decisões do autor e de outros autores para cada decisão
    porcentagem_cit = []
    # listas de tuplas com porcentagem de citações a uma decisão feita pelo próprio relator ou por outros relatores para cada decisão
    porcentagem_cit_por = []

    # casos em que não há citações são desconsiderados
    # citacoes é uma lista com de listas
    # cada lista possui a frequência de ocorrência de (1) um dado acórdão no top100
    # (2) a citações feitas por tal acórdão
    # e (3) as citações feitas a este acórdão por outros acórdãos
    for cit in citacoes:
        if cit[1] != []:
            porcentagem_cit.append(percentage_citations(cit[1], r_ids))
        if cit[2] != []:
            porcentagem_cit_por.append(percentage_citations(cit[2], r_ids))

    citacoes_top100.extend([cit[1] for cit in citacoes])

    relatores.append(r.encode('utf-8'))

    p_cit = lst_tuples_mean(porcentagem_cit)[0] if porcentagem_cit != [] else np.nan
    p_cit_por = lst_tuples_mean(porcentagem_cit_por)[0] if porcentagem_cit_por != [] else np.nan
    porc_cit_mesmo_rel.append(p_cit)
    # porcentagem de acórdãos do relator que citam as decisões dele
    porc_cit_por_mesmo_rel.append(p_cit_por)
    # gerar histograma com 3 arrays acima


citacoes_top100 = set(sum(citacoes_top100, []))

xlabel = 'Decisões'
ylabel = 'Número de ocorrências'

# relatores = []
porc_cit_mesmo_rel_top100 = []
porc_cit_alavancadas_para_top100 = []
porc_cit_alavancamento_para_top100 = []
top100_ids = sum([id.keys() for id in rel_freq_iter_dict.values()], [])
# Fazer a análise que verifica se citados estão no top 100
# Fazer a análise se acórdão é 'citado pelos' que estão no top 100
for (r, f) in sorted(rel_freq_iter_dict.items(), key=lambda i: len(i[1]), reverse=True):
    r_top100_ids = f.keys()
    citacoes = [f[k] for k in r_top100_ids]

    # frequência de ocorrência de cada acórdão no page_rank
    rel_acordaos_freq = [cit[0] for cit in citacoes]
    # invocar método que constrói histograma passando como parâmetro 'rel_acordaos_freq' e 'r.encode('utf-8')'

    title = "Número de ocorrência de decisões de %s (%s)" % (r.encode('utf-8'), sim_descr)
    file_name = "histograma_ministro_%s_%s" % (r.encode('utf-8'), sim_descr)
    generate_barplot(np.array(rel_acordaos_freq), title, xlabel, ylabel, file_name, r_top100_ids, sim_descr)

    # listas de tuplas com porcentagem de citações a decisões do autor e de outros autores para cada decisão
    porcentagem_cit = []
    # listas de tuplas com porcentagem de citações à uma decisão feitas pelo próprio relator ou por outros relatores para cada decisão
    porcentagem_cit_por = []

    # casos em que não há citações são desconsiderados
    # citacoes é uma lista com de listas
    # cada lista possui a frequência de ocorrência de (1) um dado acórdão no top100
    # (2) a citações feitas por tal acórdão
    # e (3) as citações feitas a este acórdão por outros acórdãos
    for cit in citacoes:
        if cit[1] != []:
            porcentagem_cit.append(percentage_citations(cit[1], top100_ids))
        if cit[2] != []:
            porcentagem_cit_por.append(percentage_citations(cit[2], top100_ids))

    # porcentagem de acórdãos do relator que são citados por decisões presentes no top 100
    porc_cit_mesmo_rel_top100.append(percentage_citations(r_top100_ids, citacoes_top100)[0])

    p_cit = lst_tuples_mean(porcentagem_cit)[0] if porcentagem_cit != [] else np.nan
    p_cit_por = lst_tuples_mean(porcentagem_cit_por)[0] if porcentagem_cit_por != [] else np.nan

    # porcentagem de acórdãos citados pelo relator que estão no top 100 de alguma iteração
    porc_cit_alavancadas_para_top100.append(p_cit)
    # porcentagem de acórdãos que citam os acórdãos do relator e que estão no top100
    porc_cit_alavancamento_para_top100.append(p_cit_por)


xlabel = 'Ministros'
ylabel = 'Número de decisões'
title = "Número de ocorrência de decisões por ministro \ncom repetição de ocorrência de decisões (%s)" % sim_descr
file_name = "ocorrências_de_ministros_com_repet_%s" % sim_descr
generate_barplot(np.array(relat_freqs), title, xlabel, ylabel, file_name, relatores, sim_descr)

title = "Número de ocorrências de decisões por ministro \nsem repetição de decisões (%s)" % sim_descr
file_name = "ocorrências_de_ministros_sem_repet_%s" % sim_descr
generate_barplot(np.array(relat_dec_num), title, xlabel, ylabel, file_name, relatores, sim_descr)

ylabel = 'Porcentagem'
title = "Porcentual médio de acórdãos que citam \n decisões do próprio relator [citação inversa] - (%s)" % sim_descr
file_name = "1porc_medio_cit_por_rel_%s" % sim_descr
generate_barplot(np.array(porc_cit_por_mesmo_rel), title, xlabel, ylabel, file_name, relatores, sim_descr)

title = "Porcentual médio de acórdãos citados \n por decisões do relator e que estão no top 100 \n em alguma iteração [citação direta] - (%s)" % sim_descr
file_name = "2porc_medio_cit_top100_%s" % sim_descr
generate_barplot(np.array(porc_cit_alavancadas_para_top100), title, xlabel, ylabel, file_name, relatores, sim_descr)

title = "Porcentual médio de acórdãos citados que \n pertencem ao mesmo relator [citação direta] - (%s)" % sim_descr
file_name = "3porc_medio_cit_rel_%s" % sim_descr
generate_barplot(np.array(porc_cit_mesmo_rel), title, xlabel, ylabel, file_name, relatores, sim_descr)

title = "Porcentual médio de acórdãos que citam decisões \n do próprio relator e que estão no top 100 \n em alguma iteração [citação inversa] - (%s)" % sim_descr
file_name = "4porc_medio_cit_por_top100_%s" % sim_descr
generate_barplot(np.array(porc_cit_alavancamento_para_top100), title, xlabel, ylabel, file_name, relatores, sim_descr)

title = "Porcentual de decisões do relator \n citadas por decisões no top 100 em \n alguma iteração [citação direta] - (%s)" % sim_descr
file_name = "5porc_cit_por_top100_%s" % sim_descr
generate_barplot(np.array(porc_cit_mesmo_rel_top100), title, xlabel, ylabel, file_name, relatores, sim_descr)


# DESCRIÇÃO GERAL após da por iteração
# intersecção de tuplas (ID, relator) presentes no page rank em todas as iterações
ids_relatores_intersect = reduce(set.intersection, map(set, page_rank_ids_relatores_lists))
columns = get_columns_len_tup(ids_relatores_intersect, list)
print "\nDecisões presentes em todos as iterações do page rank"
print "DECISION ID | RELATOR | VIRTUAL"
ws['A' + str(ws.max_row + 2)] = ''
ws.append(["DECISION ID", "RELATOR", "VIRTUAL"])
for pr_it in sorted(ids_relatores_intersect, key=lambda x: x[1]):
    virtual = "S" if pr_it[2] is True else "N"
    string = construct_string([pr_it[0], pr_it[1], virtual], columns)
    print string[:-2]
    ws.append([pr_it[0], pr_it[1], virtual])

print "\n"


decisions_occurance = {}
for i, pr_list in enumerate(page_ranks_iters):
    for j, pr_it in enumerate(pr_list):
        dec_id = pr_it['acordaoId']
        if dec_id in decisions_occurance:
            decisions_occurance[dec_id][-1] += 1
            decisions_occurance[dec_id][0].append((i+1,j+1))
        else:
            decisions_occurance[dec_id] = [[(i+1, j+1)], pr_it['relator'], pr_it['virtual'], 1]

frequent_decisions = filter(lambda (k, v): v[-1] > 1, decisions_occurance.items())

# ver se tem algum problema nos critérios de ordenação
frequent_decisions_sorted = sorted(frequent_decisions, key=lambda x: (x[1][-1], removed_decisions_freq[x[0]] if x[0] in removed_decisions_freq else 0, x[1][1]), reverse=True)

ws['A' + str(ws.max_row + 2)] = ''
could_be_all_iterations = 0
# precisa ver tipo exato aqui+
columns = get_columns_len_tup(frequent_decisions_sorted, None)
print "Most frequent decisions by inverse order"
print "RANK | DECISION ID | RELATOR | VIRTUAL | NUMBER OF OCCURANCES | TIMES REMOVED"
ws.append(["RANK", "DECISION ID", "RELATOR", "VIRTUAL", "NUMBER OF OCCURANCES", "TIMES REMOVED"])
# apesar de também serem armazenadas as iterações nas quais
# os IDs aparecem ache que a informação não era tão relevante
for i, (key, value) in enumerate(frequent_decisions_sorted):
    virtual = "S" if value[2] is True else "N"
    times_removed = removed_decisions_freq[key] if key in removed_decisions_freq else 0
    could_be_all_iterations += 1 if (times_removed < 10) and (times_removed + value[3]) == 10 else 0
    string = construct_string([key, value[1], virtual], columns)
    string = " {}{} |".format(i+1, " " * (4 - len(str(j)))) + string + " {} | {}".format(value[3], times_removed)
    print string
    ws.append([i+1, key, value[1], virtual, value[3], times_removed])


print "\n", collection_name
print "Number of decisions which appear on simulations: %d" % len(decisions_occurance)
print "Number of items whose frequency is higher than 1: %d" % len(frequent_decisions)
print "Number of decisions which could be in all iterations: %d" % could_be_all_iterations
ws.append(["Number of decisions which appear on simulations", len(decisions_occurance)])
ws.append(["Number of items whose frequency is higher than 1", len(frequent_decisions)])
ws.append(["Number of decisions which could be in all iterations", could_be_all_iterations])

dec_10_times = len(filter(lambda x: x[1][-1] == 10, frequent_decisions_sorted))
dec_9_times = len(filter(lambda x: x[1][-1] == 9, frequent_decisions_sorted))
dec_8_times = len(filter(lambda x: x[1][-1] == 8, frequent_decisions_sorted))

print "decisões que aparecem pelo menos 10 vezes |", dec_10_times
print "decisões que aparecem pelo menos 9 vezes |", dec_10_times + dec_9_times
print "decisões que aparecem pelo menos 8 vezes |", dec_10_times + dec_9_times + dec_8_times
ws.append(["decisões que aparecem pelo menos 10 vezes", dec_10_times])
ws.append(["decisões que aparecem pelo menos 9 vezes", dec_10_times + dec_9_times])
ws.append(["decisões que aparecem pelo menos 8 vezes", dec_10_times + dec_9_times + dec_8_times])


# capturar este número pelo script sh para extrair valor dele em script bash para concatenar finais de arquivos
print len(frequent_decisions_sorted) + 12


dec_ocur_at_leats_8_times = [acord[0] for acord in filter(lambda (k, v): v[-1] >= 8, decisions_occurance.items())]
counts_iter = []
for pr_list in page_ranks_iters:
    count = 0
    for pr_it in pr_list:
        dec_id = pr_it['acordaoId']
        if dec_id in dec_ocur_at_leats_8_times:
            count += 1

    counts_iter.append(count)


ws['A' + str(ws.max_row + 2)] = ''
ws.append(["Número médio de ocorrência por iteração das decisões que aparecem pelo menos 8 vezes",
            np.mean(np.array(counts_iter))])

ws.append(["Desvio padrão de ocorrência das decisões que aparecem pelo menos 8 vezes nas 10 iterações",
            np.std(np.array(counts_iter))])

count_decs = []
for i in range(1, 11):
    coll_rank = collection_name + "_%d" % i
    count_decs.append(db[coll_rank].count())

ws['A' + str(ws.max_row + 2)] = ''
ws.append(["Número médio de decisões na simulação", np.mean(np.array(count_decs))])
ws.append(["Desvio padrão do número de decisões na simulação", np.std(np.array(count_decs))])

ws.append(["Número médio de decisões na simulação (sem iteração 1)", np.mean(np.array(count_decs[1:]))])
ws.append(["Desvio padrão do número de decisões na simulação (sem iteração 1)", np.std(np.array(count_decs[1:]))])

wb.save(spreadsheet_name)
# -----  ver como a ordem dos acórdãos se alterna em cada iteração do page rank
# talvez adicionar informação (tupla) referente a iteração e último rank em que acórdão apareceu
# pode ser usado um dict para armazenar ambas as informações

# Exibir quantos acórdãos possuem determinada frequência - soma das frequências deve ser = 1000
